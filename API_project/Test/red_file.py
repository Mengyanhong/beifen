""""
    读取excel文件
"""

import openpyxl, sys
projectname='API_project'
print(f"{sys.argv[0].split(projectname)[0]}{projectname}/data/Excel/search_keyword.xlsx")
wb = openpyxl.load_workbook(f"{sys.argv[0].split(projectname)[0]}{projectname}/data/Excel/店铺数据准备.xlsx")
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
# sheet = wb[names[0]]
sheet = wb["店铺分类"]
# print(sheet[2])
# 获取最大行数
maxRow = sheet.max_row
print(maxRow)
# 获取最大列数
maxColumn = sheet.max_column
print(maxColumn)
# 获取当前活动表
current_sheet = wb.active
print(current_sheet)
# 获取当前活动表名称
current_name = sheet.title
print(current_name)

# 通过名字访问Cell对象, 通过value属性获取值
# a1 = sheet['A1'].value
# 通过行和列确定数据
# a12 = sheet.cell(row=1, column=2).value
# 获取列字母
column_name = openpyxl.utils.cell.get_column_letter(1)
print(column_name)
# # 将列字母转为数字, 参数忽略大小写
column_name_num = openpyxl.utils.cell.column_index_from_string(column_name)
print(column_name_num)
# # 获取一列数据, sheet.iter_rows() 获取所有的行
# """
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
# (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>)
# (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)
# (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)
# (<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>)
# """
for one_column_data in sheet.iter_rows():
    print(one_column_data[0].value)
#
# # 获取一行数据, sheet.iter_cols() 获取所有的列
# """
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
# (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>)
# (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>)
# """
# for one_row_data in sheet.iter_cols():
#     print(one_row_data[0].value, end="\t")

# print("row = {}, column = {}".format(maxRow, maxColumn))
