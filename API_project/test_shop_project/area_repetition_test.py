# -*- coding: utf-8 -*-# @Time : 2021/7/14 18:04# @Author : 孟艳红# @File : area_repetition_test.py #地区配置查重import openpyxlfrom collections import Counter #导入表格统计模块from API_project.Configs.Configuration import configuration_fileHOST = "lxcrm" #设置测试环境 test:测试环境，staging:回归环境，lxcrm:正式环境shopDivision = configuration_file(HOST).shopDivision()#实例化店铺地区并返回配置信息# print(shopDivision)# def test_shop_file(): #文件保存#     wb = openpyxl.Workbook()#     file = wb.active#     file.cell(row=1, column=1).value = "sheng_name"#     file.cell(row=1, column=2).value = "sheng_num"#     file.cell(row=1, column=3).value = "shi_name"#     file.cell(row=1, column=4).value = "shi_num"#     file.cell(row=1, column=5).value = "qu_name"#     file.cell(row=1, column=6).value = "qu_num"#     list_shenglen = len(shopDivision["normal"])#     shi_sum = 0#     qu_sum = 0#     for i in range(list_shenglen):#         file.cell(row=(i + 2), column=1).value = shopDivision["normal"][i]["NAME"]#         file.cell(row=(i + 2), column=2).value = shopDivision["normal"][i]["NUM"]#         list_shilen = len(shopDivision["normal"][i]["children"])#         for j in range(list_shilen):#             shi_sum += 1#             file.cell(row=(shi_sum + 1), column=3).value = shopDivision["normal"][i]["children"][j]["NAME"]#             file.cell(row=(shi_sum + 1), column=4).value = shopDivision["normal"][i]["children"][j]["NUM"]#             list_qulen = len(shopDivision["normal"][i]["children"][j]["children"])#             for q in range(list_qulen):#                 qu_sum += 1#                 file.cell(row=(qu_sum + 1), column=5).value = shopDivision["normal"][i]["children"][j]["children"][q][#                     "NAME"]#                 file.cell(row=(qu_sum + 1), column=6).value = shopDivision["normal"][i]["children"][j]["children"][q][#                     "NUM"]#     wb.save(r"C:/Users/admin/Desktop/店铺地区1.xlsx")#     print("保存成功")class Test_repet:    def test_repetition(self):        # print(shopDivision)        # print(shopCategory["shopDivision"]["normal"])        list_shenglen = len(shopDivision["normal"])        # print(list_shenglen)        biaoge_sheng = []        for i in range(list_shenglen):            pr_name = shopDivision["normal"][i]["NAME"]            biaoge_sheng.insert(i,pr_name)            list_shilen = len(shopDivision["normal"][i]["children"])            # print(pr_name,'\n',list_shilen)            biaoge_shi = []            for j in range(list_shilen):                city_name = shopDivision["normal"][i]["children"][j]["NAME"]                biaoge_shi.insert(j,city_name)                list_qulen = len(shopDivision["normal"][i]["children"][j]["children"])                # print(list_qulen)                biaoge_qu = []                for q in range(list_qulen):                    qu_name = shopDivision["normal"][i]["children"][j]["children"][q]["NAME"]                    biaoge_qu.insert(q,qu_name)                # print(biaoge_qu)                b_qu = dict(Counter(biaoge_qu))                # print(b_qu)                b__qu = {key:value for key,value in b_qu.items() if value > 1}                # print(b__qu)                if b__qu != {}:                    print(pr_name,city_name)                    print(b__qu)                    assert b__qu == {}                else:                    assert b__qu == {}            b_shi = dict(Counter(biaoge_shi))            # print(b_shi)            if [key for key,value in b_shi.items() if value > 1] != []:                print([key for key, value in b_shi.items() if value > 1])                print(pr_name)                assert [key for key,value in b_shi.items() if value > 1] == []            else:                assert [key for key, value in b_shi.items() if value > 1] == []        # print(biaoge_sheng)        b_sheng = dict(Counter(biaoge_sheng))        print(b_sheng)        if {key:value for key,value in b_sheng.items()if value > 1} != {}:            print({key:value for key,value in b_sheng.items()if value > 1})            assert {key:value for key,value in b_sheng.items()if value > 1} == {}        else:            assert {key: value for key, value in b_sheng.items() if value > 1} == {}# if __name__ == '__main__':