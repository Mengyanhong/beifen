import openpyxl, sys, json
projectname='API_project'
print(f"{sys.argv[0].split(projectname)[0]}{projectname}/data/Excel/search_keyword.xlsx")
class Excel_Files:
    def __init__(self, file_name='访客识别-用例.xlsx', sheel='中企接口', projectname='home'):
        '''
        :param file_name: 用例文件
        :param sheel: 工作表
        '''
        self.file_name = file_name
        self.sheel = sheel
        self.file = openpyxl.load_workbook(
            f"{sys.path[0].split(projectname)[0]}{projectname}/API_project/data/Excel/{self.file_name}")  # 打开用例文件
        self.sheet_header = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']  # 创建文件表头列表
        self.sheel_file = self.file[self.sheel]  # 打开用例文件内容

    def excel_index(self, index):
        '''
        :param index: 文件表头
        :return: 表头索引
        '''
        row_list = []  # 创建case名称标题列表
        for i in self.sheel_file[1]:  # 获取case列表标题
            row_list.append(i.value)  # 将case标题添加到名称标题列表，追加模式
        case_index = row_list.index(index)  # 获取用例索引
        return case_index

    def open_file_rows(self, index):
        row_list = []  # 创建case名称标题列表
        for i in self.sheel_file[1]:  # 获取case列表标题
            row_list.append(i.value)  # 将case标题添加到名称标题列表，追加模式
        files_list = []
        # print(list(self.sheel_file.rows))
        for i in list(self.sheel_file.rows):
            if i[row_list.index(index)].value == None:
                continue
            elif i[row_list.index(index)].value == index:
                continue
            else:
                files_list.append(i[row_list.index(index)].value)
        # print(self.sheel_file.cell("A").value)  # 获取用例标题
        return files_list

    def open_case(self, request_name='入参', response_name='出餐', Sizer_name='用例标题', Sizer_name_v='前置条件', Sizer_value=[]):
        '''
        :param request_name: 入参表头
        :param response_name: 出参表头
        :param Sizer_name: 用例标题表头
        :param Sizer_value: 用例标题
        :return: 用例
        '''
        case_list = []
        Sizer_name_len = len(
            self.sheel_file[self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(Sizer_name)]])
        # print(
        #     len(self.sheel_file[self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)]])) #打印用例数量含空用例
        for i in range(Sizer_name_len):  # 通过用例索引循环全部用例
            Sizer_name_value = (self.sheel_file[
                str(self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(Sizer_name)]) + str(
                    i + 1)]).value  # 获取用例标题
            Sizer_name_valu = (self.sheel_file[
                str(self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(Sizer_name_v)]) + str(
                    i + 1)]).value  # 获取用例标题

            if Sizer_name_value is not None and Sizer_name_value != Sizer_name:  # 判断用例标题不为空 跳过第一行标题
                response_value = (self.sheel_file[
                    self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(response_name)] + str(
                        i + 1)]).value  # 通过响应索获取响应内容
                request_value = (self.sheel_file[
                    self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)] + str(
                        i + 1)]).value  # 通过请求索获取请求内容
                #     continue
                if isinstance(request_value, str):  # 首先判断变量是否为字符串
                    try:
                        congig_text = json.loads(request_value)  #
                    except ValueError:
                        request_value = request_value
                    else:
                        request_value = congig_text
                if isinstance(response_value, str):  # 首先判断变量是否为字符串
                    try:
                        congig_text = json.loads(response_value)  #
                    except ValueError:
                        response_value = response_value
                    else:
                        response_value = congig_text
                if Sizer_value != []:  # 判断用例筛选不为空
                    if Sizer_name_value in Sizer_value:  # 判断用例符合筛选内容
                        case_list.append((request_value, response_value, Sizer_name_valu))  # 将用例入参和出餐写入用例表格
                else:
                    case_list.append((request_value, response_value, Sizer_name_valu))  # 若不进行筛选 直接将用例入参和出餐写入用例表格
        return case_list

    def open_case2(self, request_name='入参', response_name='出餐', Sizer_name='用例标题', Sizer_name_v='前置条件', Sizer_value=[]):
        '''
        :param request_name: 入参表头
        :param response_name: 出参表头
        :param Sizer_name: 用例标题表头
        :param Sizer_value: 用例标题
        :return: 用例
        '''
        case_list = []
        Sizer_name_len = len(
            self.sheel_file[self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)]])
        # print(
        #     len(self.sheel_file[self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)]])) #打印用例数量含空用例
        for i in range(Sizer_name_len):  # 通过用例索引循环全部用例
            Sizer_name_value = (self.sheel_file[
                str(self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)]) + str(
                    i + 1)]).value  # 获取用例标题
            Sizer_name_valu = (self.sheel_file[
                str(self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(response_name)]) + str(
                    i + 1)]).value  # 获取用例标题

            if Sizer_name_value is not None and Sizer_name_value != Sizer_name:  # 判断用例标题不为空 跳过第一行标题
                response_value = (self.sheel_file[
                    self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(response_name)] + str(
                        i + 1)]).value  # 通过响应索获取响应内容
                request_value = (self.sheel_file[
                    self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)] + str(
                        i + 1)]).value  # 通过请求索获取请求内容
                #     continue
                if isinstance(request_value, str):  # 首先判断变量是否为字符串
                    try:
                        congig_text = json.loads(request_value)  #
                    except ValueError:
                        request_value = request_value
                    else:
                        request_value = congig_text
                if isinstance(response_value, str):  # 首先判断变量是否为字符串
                    try:
                        congig_text = json.loads(response_value)  #
                    except ValueError:
                        response_value = response_value
                    else:
                        response_value = congig_text
                if Sizer_value != []:  # 判断用例筛选不为空
                    if Sizer_name_value in Sizer_value:  # 判断用例符合筛选内容
                        case_list.append((request_value, response_value))  # 将用例入参和出餐写入用例表格
                else:
                    case_list.append((request_value, response_value))  # 若不进行筛选 直接将用例入参和出餐写入用例表格
        return case_list

    def open_case3(self, *Sizer_row_value, request_name='入参', response_name='出餐', Sizer_column_value=[]):
        '''
        :param request_name: 入参表头
        :param response_name: 出参表头
        :param Sizer_name: 用例标题表头
        :param Sizer_value: 用例标题
        :return: 用例
        '''
        # print(Sizer_row_value)
        case_list = []
        max_rows = self.sheel_file.max_row + 1
        # print(max_rows)
        request_value_test = None
        # print(
        #     len(self.sheel_file[self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)]])) #打印用例数量含空用例
        for i in range(2,max_rows):  # 通过用例索引循环全部用例
            request_value = (self.sheel_file[
                self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(request_name)] + str(
                    i)]).value  # 通过请求索获取请求内容
            if request_value is not None:
                request_value_test = request_value
            # if isinstance(request_value_test, str):  # 首先判断变量是否为字符串
            #     try:
            #         request_value_test = json.loads(request_value_test)  #
            #     except ValueError:
            #         request_value_test = request_value_test
            # if isinstance(response_value, str):  # 首先判断变量是否为字符串
            #     try:
            #         response_value = json.loads(response_value)  #
            #     except ValueError:
            #         response_value = response_value

            if Sizer_row_value:  # 判断用例筛选不为空
                if request_value_test in Sizer_row_value:  # 判断用例符合筛选内容
                    response_value = (self.sheel_file[
                        self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(response_name)] + str(
                            i)]).value  # 通过响应索获取响应内容
                    if response_value is not None:
                        case_list.append({"categoryL1_value": request_value_test, "categoryL2_value": response_value})
                        # case_list.append((request_value_test, response_value))  # 将用例入参和出餐写入用例表格
                else:
                    break
            else:
                response_value = (self.sheel_file[
                    self.sheet_header[Excel_Files(self.file_name, self.sheel).excel_index(response_name)] + str(
                        i)]).value  # 通过响应索获取响应内容
                if response_value is not None:
                    case_list.append({"categoryL1_value": request_value_test, "categoryL2_value": response_value})
                # case_list.append((request_value, response_value))  # 若不进行筛选 直接将用例入参和出餐写入用例表格
            # print(case_list)
        return case_list

if __name__ == '__main__':
    print(sys.path[0])
    # EXcel_file = Excel_Files(file_name="联系方式渠道配置.xlsx", sheel="联系方式渠道配置")
    # EXcel_file = Excel_Files(file_name="search_keyword.xlsx", sheel="search_keyword")
    excel_file = Excel_Files(file_name="店铺数据准备.xlsx", sheel="店铺分类")
    file = excel_file.open_case3("美食", request_name="一级分类", response_name="二级分类")

    # EXcel_file = Excel_Files(file_name="联系方式渠道配置test.xlsx", sheel="联系方式渠道配置").open_case2(request_name="name", response_name="value",
    #                                                                        Sizer_name="name")[90:100]
    print(file)
    # print(EXcel_file.excel_index(index="pid"))
    # # print(EXcel_file.open_file_rows("name"))
    # print(EXcel_file.open_file_rows("entName"))
