# -*- coding: utf-8 -*-
# @Time : 2021/10/15 11:58
# @Author : 孟艳红
# @File : sql.py

from Visitors_identify.Configs.Pymysql import to_pymysql
import random
import pytest,os
host = 'staging' #设置测试环境 test:测试环境，staging:回归环境，lxcrm:正式环境
from openpyxl import load_workbook
# identify_excel=load_workbook(r"C:\Users\admin\Desktop\site_visitor_identify.xlsx") #打开数据表格
# #获取sheet：
# identify_table = identify_excel["site_visitor_identify"]   #通过表名获取
# row_list = [] #创建case标题列表
# for i in identify_table[1]: #获取case列表标题
#     row_list.append(i.value) #将case标题写入列表
# rows=identify_table.max_row   #获取行数
# cols=identify_table.max_column    #获取列数
# for i in range(2,rows+1):
#     print(table.cell(row=i,column=site_id).value,str(table.cell(row=i,column=latest_visit_time).value).split(".")[0]) #获取单元格值：data=table.cell(row=row,column=col).value  #获取表格内容，是从第row行第col列开始的，注意不要丢掉 .value

site_id = 1099
oid = 5005648
class sql:
    def site_visitor_identify_sql_insert(self):
        pymysql = to_pymysql(host).pymysql()
        cursor = pymysql.cursor()
        identify_excel = load_workbook(r"C:\Users\admin\Desktop\site_visitor_identify.xlsx")  # 打开数据表格
        # 获取sheet：
        identify_table = identify_excel["site_visitor_identify"]  # 通过表名获取
        row_list = []  # 创建case标题列表
        for i in identify_table[1]:  # 获取case列表标题
            row_list.append(i.value)  # 将case标题写入列表
        rows = identify_table.max_row  # 获取行数
        sums=0
        for i in range(45, rows + 1): #循环列表数据
            sums+=1
            if sums == 1000:
                break
            else:
                ip = row_list.index("ip") + 1  # 获取用例索引
                ip_user = row_list.index("ip_user") + 1  # 获取用例索引
                pid = row_list.index("pid") + 1  # 获取用例索引
                first_visit_platform = row_list.index("first_visit_platform") + 1  # 获取用例索引
                first_visit_time = row_list.index("first_visit_time") + 1  # 获取用例索引
                latest_visit_time = row_list.index("latest_visit_time") + 1  # 获取用例索引
                ips = identify_table.cell(row=i, column=ip).value  # 获取用例值
                ip_users = identify_table.cell(row=i, column=ip_user).value  # 获取用例值
                pids = identify_table.cell(row=i, column=pid).value  # 获取用例值
                first_visit_platforms = identify_table.cell(row=i, column=first_visit_platform).value  # 获取用例值
                first_visit_times = str(identify_table.cell(row=i, column=first_visit_time).value).split(".")[0]
                latest_visit_times = str(identify_table.cell(row=i, column=latest_visit_time).value).split(".")[0]   # 获取用例值
                print(site_id,str(ips),ip_users,pids,str(first_visit_platforms),first_visit_times,str(latest_visit_times))

                cursor.execute(
                    f"""insert into site_visitor_identify (site_id,ip,ip_user,pid,first_visit_platform,first_visit_time,latest_visit_time,oid) values
                     ({site_id},"{ips}","{ip_users}","{pids}",{first_visit_platforms},"{first_visit_times}","{latest_visit_times}",{oid})""") #插入测试数据
                pymysql.commit()
        cursor.close()
        pymysql.close()
    def site_visitor_sql_insert(self):
        pymysql = to_pymysql(host).pymysql()
        cursor = pymysql.cursor()
        visitor_excel = load_workbook(r"C:\Users\admin\Desktop\site_visitors.xlsx")  # 打开数据表格
        # 获取sheet：
        visitor_table = visitor_excel["site_visitors"]  # 通过表名获取
        row_list = []  # 创建case标题列表
        for i in visitor_table[1]:  # 获取case列表标题
            row_list.append(i.value)  # 将case标题写入列表
        rows = visitor_table.max_row  # 获取行数
        print(row_list)
        sums=0
        for i in range(2, rows + 1): #循环列表数据
            sums+=1
            if sums == 100:
                break
            else:
                # 获取用例索引
                ip_user = row_list.index("ip_user") + 1  # 获取用例索引
                pid = row_list.index("pid") + 1  # 获取用例索引
                visit_count = row_list.index("visit_count") + 1  # 获取用例索引
                visit_total_secs = row_list.index("visit_total_secs") + 1  # 获取用例索引
                first_visit_time = row_list.index("first_visit_time") + 1  # 获取用例索引
                latest_visit_time = row_list.index("latest_visit_time") + 1  # 获取用例索引
                # 获取用例参数
                ip_users = visitor_table.cell(row=i, column=ip_user).value  # 获取用例值
                pids = visitor_table.cell(row=i, column=pid).value  # 获取用例值
                visit_counts = visitor_table.cell(row=i, column=visit_count).value  # 获取用例值
                visit_total_secss = visitor_table.cell(row=i, column=visit_total_secs).value  # 获取用例值
                first_visit_times = str(visitor_table.cell(row=i, column=first_visit_time).value).split(".")[0] # 获取用例值
                latest_visit_times = str(visitor_table.cell(row=i, column=latest_visit_time).value).split(".")[0]   # 获取用例值
                visitor_id = random.randint(100000,200000)
                # print(site_id,ip_users,pids,first_visit_times,str(latest_visit_times))

                cursor.execute(
                    f"""insert into site_visitors (site_id,ip_user,pid,visit_count,visit_total_secs,first_visit_time,latest_visit_time,visitor_id) values
                     ({site_id},"{ip_users}","{pids}",{visit_counts},{visit_total_secss},"{first_visit_times}","{latest_visit_times}","{visitor_id}")""") #插入测试数据
                pymysql.commit()
        cursor.close()
        pymysql.close()

if __name__ == '__main__':
    # sql().site_visitor_identify_sql_insert()
    sql().site_visitor_sql_insert()