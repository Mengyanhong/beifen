from operator import index
import openpyxl
import time
import requests
import numpy as np
from elasticsearch import Elasticsearch
from openpyxl import load_workbook


# 获取文件：
print('请先把文件转化为Excel后缀为 .xlsx，不要改文件名')
filename = input('请复制文件名称：')
filepath = r"C:/Users/admin/Desktop/桌面文件/测试脚本/python文件/" + filename + ".xlsx"  # 获取路径
file = load_workbook(filepath)  # 获取文件内容
# 获取sheet：
# sheet_names = file.get_sheet_names()#若知道sheet名称可直接输入，file.get_sheet_names(sheet1)
# sheet = file.get_sheet_by_name(sheet_names[index])  # index为0为第一张表,==file.active
sheet = file.active
rows = sheet.max_row
print(rows)  # 获取行数
cols = sheet.max_column
print(cols)  # 获取列数
# file_Data=sheet.cell(row=rows,column=cols).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
column_name = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
shopcategory=[{"name":"10","value":"美食","sub":[{"name":"112","value":"小吃快餐"},{"name":"110","value":"火锅"},{"name":"2714","value":"水果生鲜"},{"name":"508","value":"烧烤"},{"name":"118","value":"其他美食"},{"name":"34236","value":"饮品店"},{"name":"102","value":"川菜"},{"name":"117","value":"面包/饮品"},{"name":"33759","value":"食品保健"},{"name":"215","value":"面馆"},{"name":"116","value":"西餐"},{"name":"101","value":"本帮江浙菜"},{"name":"114","value":"韩国料理"},{"name":"132","value":"咖啡厅"},{"name":"106","value":"东北菜"},{"name":"25474","value":"农家菜"},{"name":"251","value":"海鲜"},{"name":"103","value":"潮粤菜"},{"name":"104","value":"湘菜"},{"name":"1783","value":"家常菜"},{"name":"113","value":"日本菜"},{"name":"34284","value":"特色菜"},{"name":"1338","value":"私房菜"},{"name":"219","value":"小龙虾"},{"name":"311","value":"北京菜"},{"name":"111","value":"自助餐"},{"name":"3243","value":"新疆菜"},{"name":"6743","value":"云贵菜"},{"name":"26481","value":"西北菜"},{"name":"250","value":"创意菜"},{"name":"115","value":"东南亚菜"},{"name":"1817","value":"粉面馆"},{"name":"246","value":"湖北菜"},{"name":"109","value":"素食"},{"name":"1959","value":"粥粉面"},{"name":"27686","value":"海南菜"},{"name":"34059","value":"福建菜"},{"name":"1080","value":"豫菜"},{"name":"107","value":"台湾菜"},{"name":"249","value":"闽菜"},{"name":"497","value":"天津菜"}]}]
print(len(shopcategory))
def industryL1():
    sum = 0
    for i in range(2, rows + 2):
        file_Data = sheet[column_name[2] + str(i)].value
        file_Data_1 = sheet[column_name[3] + str(i)].value
        if file_Data != None:
            url = "https://skb-test.weiwenjia.com/api_skb/v1/advanced_search"
            headers = {"Authorization": "Token token=c226f5c5b96665b6aae7ec832e997408",
                       "crm_platform_type": "lixiaoyun"}
            json = {
                "hasSyncClue": 0,
                "hasSyncRobot": 0,
                "hasUnfolded": 0,
                "sortBy": 0,
                "condition":
                    {
                        "cn": "composite",
                        "cr": "MUST",
                        "cv":
                            [
                                {
                                    "cn": "industry",
                                    "cr": "IN",
                                    "cv": {"firstIndustry": [file_Data],
                                           "secondIndustry": []}
                                }
                            ]
                    },
                "pagesize": 10,
                "page": 1,
                "templateType": 0,
                "templateName": "",
                "userClick": 1
            }
            response = requests.post(url, headers=headers, json=json)
            response = response.json()
            print(response)
            print(file_Data)
            for j in range(1):
                res_pid = response["data"]["items"][j]["id"]  # 获取pid
                res_industryL1 = response["data"]["items"][j]["industry"]
                # 调用es接口查询接口信息
                es_client = Elasticsearch('es-cn-nif1oiv5w0009di0f.public.elasticsearch.aliyuncs.com:9200',
                                          http_auth=('lihexiang', 'Aa123456'))
                result = es_client.get(index="company_info_prod", id=res_pid)['_source']
                result_industry = result["firstIndustry"]  # 获取二级行业值
                if result_industry != file_Data:
                    print("一级行业搜索错误")
                print(res_industryL1, file_Data_1)
                if res_industryL1 != file_Data_1:
                    print("一级行业搜索错误")
                time.sleep(0.5)
        else:
            sum += 1
            if sum > 2:
                print("测试结束测试数据调取完毕")
                break

            # print(result["firstIndustry"])
            # print(result["thirdIndustry"])


def industryL2():
    sum_2 = 0
    for i in range(2, rows + 2):
        file_Data = sheet[column_name[6] + str(i)].value
        if file_Data != None:
            url = "https://skb-test.weiwenjia.com/api_skb/v1/advanced_search"
            headers = {"Authorization": "Token token=c226f5c5b96665b6aae7ec832e997408",
                       "crm_platform_type": "lixiaoyun"}
            json = {
                "hasSyncClue": 0,
                "hasSyncRobot": 0,
                "hasUnfolded": 0,
                "sortBy": 0,
                "condition":
                    {
                        "cn": "composite",
                        "cr": "MUST",
                        "cv":
                            [
                                {
                                    "cn": "industry",
                                    "cr": "IN",
                                    "cv": {"firstIndustry": [],
                                           "secondIndustry": [file_Data]}
                                }
                            ]
                    },
                "pagesize": 10,
                "page": 1,
                "templateType": 0,
                "templateName": "",
                "userClick": 1
            }
            response = requests.post(url, headers=headers, json=json)
            response = response.json()

            for j in range(1):
                res_pid = response["data"]["items"][j]["id"]  # 获取pid
                # 调用es接口查询接口信息
                es_client = Elasticsearch('es-cn-nif1oiv5w0009di0f.public.elasticsearch.aliyuncs.com:9200',
                                          http_auth=('mengyanhong', 'Aa123456'))
                result = es_client.get(index="company_info_prod", id=res_pid)['_source']  # 执行搜索

                result_industry_2 = result["secondIndustry"]  # 获取二级行业数据
                result_industry_3 = [int(x) for x in result_industry_2]  # 将二级行业数据转换为int
                arr = np.array(result_industry_3)  # 将二级行业数据先转化为numpy.ndarray
                # if file_Data<10:
                #     file_Data_1=str(0)+str(file_Data)
                # else:
                #     file_Data_1=str(file_Data)
                # print(file_Data_1)
                if (arr == file_Data).any() == False:  # 判断二级行业数据中是否包含搜索条件False不包含True包含
                    print("二级行业搜索错误")
            if i == rows:
                print("测试结束测试数据调取完毕")


if __name__ == "__main__":
    industryL2()
