import pymysql
import openpyxl
import requests
import threading

wb = openpyxl.Workbook()
wl = wb.create_sheet("1")
ws = wb.active
# ws.cell(row=1, column=1).value = "时间"
# ws.cell(row=1, column=2).value = "评论"
# ws.cell(row=1, column=3).value = "长度"
# # wb.save("Python.xlsx")
# print("保存成功")
# 打开数据库连接
biaoge = []
biao = ['1']
connect = pymysql.connect(host="rdscbq34656z0ix59br0.mysql.rds.aliyuncs.com", port=3306, user="yxy_skb_test",
                          password="Yxyskb#ae964fd1", database="yxy_skb_test")
cursor = connect.cursor()  # 获取游标
cursor.execute("select pid from recommendation_task_result where task_id = 149")  # 执行语句
cu = cursor.fetchone()  # 获取第一个数据（字典）
r1 = cursor.fetchall()  # 获取全部数据（字典）# 结果是元组,fetchone()获取查询结果
len = len(r1)
# list_r1 = list(r1)
# print(len, list_r1)
for i in range(len):
    cl = str(r1[i][0])  # 从字典内取第一个字符串(str)
    ws.cell(row=(i + 2), column=1).value = cl
    biaoge.insert(i, cl)
    # print (type(cl))
    # print (cl)
# tupl = tuple(biaoge)
# # ws.cell(row=2, column=2).value = tupl
# print(list(tupl))
# dic = dict(zip(biaoge,biao))
# ws.cell(row=2, column=3).value = str(biaoge)
# wb.save("Python1.xlsx")
wb.save(r"D:\桌面文件\Python1.xlsx")
# print(biaoge)
# biaoge.save("biao.xlsx")
# print(list(cu))
cursor.close()
connect.close()


# def test_post():
#     url = "https://skb-test.weiwenjia.com/api_skb/v1/clues/sync"
#     headers = {
#         'Authorization': 'Token token=8b07d5bc522096e6bdabfc68fc8435e8',
#         'crm_platform_type': 'lixiaoyun'
#     }
#     json = {
#         "keyword": "",
#         "filter":
#             {
#                 "location": [],
#                 "industryshort": [],
#                 "secindustryshort": [],
#                 "registercapital": [],
#                 "establishment": [],
#                 "entstatus": [],
#                 "contact": [],
#                 "sortBy": "0",
#                 "companysource": [],
#                 "enttype": [0],
#                 "employees": [0],
#                 "hasrecruit": "0",
#                 "hassem": "0",
#                 "haswebsite": "0",
#                 "hastrademark": "0",
#                 "haspatent": "0",
#                 "hastender": "0",
#                 "haswechataccnt": "0",
#                 "filterUnfold": 0,
#                 "filterSync": 0,
#                 "filterSyncRobot": 0,
#                 "hasBuildingCert": "0",
#                 "isHighTech": "0",
#                 "hasFinanceInfo": "0",
#                 "hasAbnormalInfo": "0"
#             },
#         "scope": "companyname",
#         "pids": biaoge,
#         "way": "search_list",
#         "from": "syncClue",
#         "useQuota": "true"
#     }
#     r = requests.post(url, headers=headers, json=json)
#     # print(json["pids"])
#     print(r.text)
#
#
# def test_shield():
#     url = "https://shield.lixiaoskb.com/api/dq/select"
#     hraders = {
#         "authorization": "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJfaWQiOiI1Zjk2YmFhYjE1ZmJhNTJmOGM2ZGE5MGYiLCJyb2xlIjoiYWRtaW4iLCJpYXQiOjE2MTA2ODE5OTksImV4cCI6MTYxMDc2ODM5OX0.VVpvifYVsf5pV0aHFLsIi-Iqc3hVt1Q6c1Xx0pgsodQ",
#         "origin": "https://shield.lixiaoskb.com",
#         "referer": "https://shield.lixiaoskb.com/dq",
#         "sec-ch-ua": '"Google Chrome";v="87", " Not;A Brand";v="99", "Chromium";v="87"'
#     }
#     json = {"dbName": "enterprise", "collection": "EnterpriseBaseInfo", "sort": "", "filter": "{}", "projection": "",
#             "sortType": -1, "limit": 10}
#     response = requests.post(url, json=json, headers=hraders)  # ,verify=False取消证书验证
#     print(response.json())
#
#
#
# if __name__ == "__main__":
#     # threads = [threading.Thread(target=test_shield),
#     #            threading.Thread(target=test_post)]
#     # for t in threads:
#     #     # 启动线程
#     #     t.start()
#     test_shield()

