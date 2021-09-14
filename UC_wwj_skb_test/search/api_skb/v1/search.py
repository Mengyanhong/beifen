from operator import index
import openpyxl
import time
import requests
import numpy as np
from elasticsearch import Elasticsearch
from openpyxl import load_workbook

# 获取文件：
print('请先把文件转化为Excel后缀为 .xlsx，不要改文件名')
filename = input('请复制文件名称：')
filepath = r"C:/Users/admin/Desktop/桌面文件/测试脚本/python文件/" + filename + ".xlsx"  # 获取路径
file = load_workbook(filepath)  # 获取文件内容
# 获取sheet：
# sheet_names = file.get_sheet_names()#若知道sheet名称可直接输入，file.get_sheet_names(sheet1)
# sheet = file.get_sheet_by_name(sheet_names[index])  # index为0为第一张表,==file.active
sheet = file.active
rows = sheet.max_row
print(rows)  # 获取行数
cols = sheet.max_column
print(cols)  # 获取列数
# file_Data=sheet.cell(row=rows,column=cols).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
column_name = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]


def industryL1():
    sum = 0
    for i in range(2, rows + 2):
        file_Data = sheet[column_name[2] + str(i)].value
        file_Data_1 = sheet[column_name[3] + str(i)].value
        if file_Data != None:
            url = "https://skb-test.weiwenjia.com/api_skb/v1/advanced_search"
            headers = {"Authorization": "Token token=c226f5c5b96665b6aae7ec832e997408",
                       "crm_platform_type": "lixiaoyun"}
            json = {
                "hasSyncClue": 0,
                "hasSyncRobot": 0,
                "hasUnfolded": 0,
                "sortBy": 0,
                "condition":
                    {
                        "cn": "composite",
                        "cr": "MUST",
                        "cv":
                            [
                                {
                                    "cn": "industry",
                                    "cr": "IN",
                                    "cv": {"firstIndustry": [file_Data],
                                           "secondIndustry": []}
                                }
                            ]
                    },
                "pagesize": 10,
                "page": 1,
                "templateType": 0,
                "templateName": "",
                "userClick": 1
            }
            response = requests.post(url, headers=headers, json=json)
            response = response.json()
            print(response)
            print(file_Data)
            for j in range(1):
                res_pid = response["data"]["items"][j]["id"]  # 获取pid
                res_industryL1 = response["data"]["items"][j]["industry"]
                # 调用es接口查询接口信息
                es_client = Elasticsearch('es-cn-nif1oiv5w0009di0f.public.elasticsearch.aliyuncs.com:9200',
                                          http_auth=('lihexiang', 'Aa123456'))
                result = es_client.get(index="company_info_prod", id=res_pid)['_source']
                result_industry = result["firstIndustry"]  # 获取二级行业值
                if result_industry != file_Data:
                    print("一级行业搜索错误")
                print(res_industryL1, file_Data_1)
                if res_industryL1 != file_Data_1:
                    print("一级行业搜索错误")
                time.sleep(0.5)
        else:
            sum += 1
            if sum > 2:
                print("测试结束测试数据调取完毕")
                break

            # print(result["firstIndustry"])
            # print(result["thirdIndustry"])


def industryL2():
    sum_2 = 0
    for i in range(2, rows + 2):
        file_Data = sheet[column_name[6] + str(i)].value
        if file_Data != None:
            url = "https://skb-test.weiwenjia.com/api_skb/v1/advanced_search"
            headers = {"Authorization": "Token token=c226f5c5b96665b6aae7ec832e997408",
                       "crm_platform_type": "lixiaoyun"}
            json = {
                "hasSyncClue": 0,
                "hasSyncRobot": 0,
                "hasUnfolded": 0,
                "sortBy": 0,
                "condition":
                    {
                        "cn": "composite",
                        "cr": "MUST",
                        "cv":
                            [
                                {
                                    "cn": "industry",
                                    "cr": "IN",
                                    "cv": {"firstIndustry": [],
                                           "secondIndustry": [file_Data]}
                                }
                            ]
                    },
                "pagesize": 10,
                "page": 1,
                "templateType": 0,
                "templateName": "",
                "userClick": 1
            }
            response = requests.post(url, headers=headers, json=json)
            response = response.json()
            for j in range(1):
                res_pid = response["data"]["items"][j]["id"]  # 获取pid
                # 调用es接口查询接口信息
                es_client = Elasticsearch('es-cn-nif1oiv5w0009di0f.public.elasticsearch.aliyuncs.com:9200',
                                          http_auth=('lihexiang', 'Aa123456'))
                result = es_client.get(index="company_info_prod", id=res_pid)['_source'] # 执行搜索
                result_industry_2 = result["secondIndustry"] # 获取二级行业数据
                result_industry_3 = [int(x) for x in result_industry_2] # 将二级行业数据转换为int
                arr=np.array(result_industry_3)# 将二级行业数据先转化为numpy.ndarray
                # if file_Data<10:
                #     file_Data_1=str(0)+str(file_Data)
                # else:
                #     file_Data_1=str(file_Data)
                # print(file_Data_1)
                if (arr == file_Data).any()==False:#判断二级行业数据中是否包含搜索条件
                    print("二级行业搜索错误")
                # else:
                #     print(file_Data,arr)
            if i == rows:
                print("测试结束测试数据调取完毕")
                # print((arr == str(file_Data_1)).any())
        # else:
        #     sum_2 += 1
        #     if sum_2 > 2:
        #         print("测试结束测试数据调取完毕")
        #         break


def gaojisearch():
    url = "https://skb-test.weiwenjia.com/api_skb/v1/advanced_search"
    headers = {"Authorization": "Token token=c226f5c5b96665b6aae7ec832e997408",
               "crm_platform_type": "lixiaoyun"
               }
    json = {
        "hasSyncClue": 0,
        "hasSyncRobot": 0,
        "hasUnfolded": 0,
        "sortBy": 0,
        "condition":
            {
                "cn": "composite",
                "cr": "MUST",
                "cv":
                    [{
                        "cn": "industry",
                        "cr": "IN",
                        "cv":
                            {
                                "firstIndustry": ["A"],
                                "secondIndustry": []
                            }
                        # "id":"16109362808701"
                    }
                    ]
            },
        "page": 1,
        "pagesize": 10,
        "templateType": 0,
        "templateName": "",
        "userClick": 1
    }
    response = requests.post(url, headers=headers, json=json)
    print(response.json())
    print(response.text)


if __name__ == "__main__":
    industryL2()
