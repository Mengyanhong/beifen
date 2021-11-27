from collections import Counter
# a = ["REGION_8194","REGION_72227","REGION_8194"]
# b = dict(Counter(a))
# print([key for key,value in b.items() if value > 1])


# import requests
# from elasticsearch import Elasticsearch
#
# es_client = Elasticsearch('es-cn-nif1oiv5w0009di0f.public.elasticsearch.aliyuncs.com:9200',
#                           http_auth=('lihexiang', 'Aa123456'))
# result = es_client.get(index="shop_info_prod", id="126761135")['_source']
# print(type(result))
# print(result)
# print(result["city"])
# print(result["province"])
# print(result["district"][0])
import openpyxl
import requests
import numpy as np
from elasticsearch import Elasticsearch

# list_shenglen = len(shopDivision["normal"][0]["NUM"])
# list_shilen = len(shopDivision["normal"][0]["children"][0]["NUM"])
# list_qulen = len(shopDivision["normal"][0]["children"][0]["children"][0]["NUM"])
# print(list_shenglen,list_shilen,list_qulen)

def shop_file():
    wb = openpyxl.Workbook()
    file = wb.active
    file.cell(row=1, column=1).value = "sheng_name"
    file.cell(row=1, column=2).value = "sheng_num"
    file.cell(row=1, column=3).value = "shi_name"
    file.cell(row=1, column=4).value = "shi_num"
    file.cell(row=1, column=5).value = "qu_name"
    file.cell(row=1, column=6).value = "qu_num"
    list_shenglen = len(shopDivision["normal"])
    shi_sum = 0
    qu_sum = 0
    URL = "https://skb.weiwenjia.com/api_skb/v1/shop_search"
    headers = {"Authorization": "Token token=1ceb18ed4a12c77bdb283ee3e9d0ab98",
               "crm_platform_type": "lixiaoyun"}
    for i in range(list_shenglen):
        file.cell(row=(i + 2), column=1).value = shopDivision["normal"][i]["NAME"]
        file.cell(row=(i + 2), column=2).value = shopDivision["normal"][i]["NUM"]
        list_shilen = len(shopDivision["normal"][i]["children"])
        for j in range(list_shilen):
            shi_sum += 1
            file.cell(row=(shi_sum + 1), column=3).value = shopDivision["normal"][i]["children"][j]["NAME"]
            file.cell(row=(shi_sum + 1), column=4).value = shopDivision["normal"][i]["children"][j]["NUM"]
            list_qulen = len(shopDivision["normal"][i]["children"][j]["children"])
            for q in range(list_qulen):
                qu_sum += 1
                file.cell(row=(qu_sum + 1), column=5).value = shopDivision["normal"][i]["children"][j]["children"][q][
                    "NAME"]
                file.cell(row=(qu_sum + 1), column=6).value = shopDivision["normal"][i]["children"][j]["children"][q][
                    "NUM"]
    wb.save(r"C:/Users/admin/Desktop/桌面文件/店铺地区1.xlsx")
    print("保存成功")
def shop_search():
    list_shenglen = len(shopDivision["normal"])
    biaoge_sheng = []
    for i in range(list_shenglen):
        pr_name = shopDivision["normal"][i]["NAME"]
        biaoge_sheng.insert(i,pr_name)
        list_shilen = len(shopDivision["normal"][i]["children"])
        biaoge_shi = []
        for j in range(list_shilen):
            city_name = shopDivision["normal"][i]["children"][j]["NAME"]
            biaoge_shi.insert(j,city_name)
            list_qulen = len(shopDivision["normal"][i]["children"][j]["children"])
            biaoge_qu = []
            for q in range(list_qulen):
                qu_name = shopDivision["normal"][i]["children"][j]["children"][q]["NAME"]
                biaoge_qu.insert(q,qu_name)
            b_qu = dict(Counter(biaoge_qu))
            b__qu = {key:value for key,value in b_qu.items() if value > 1}

            if b__qu != {}:
                print({key: value for key, value in b_qu.items() if value > 1})
                print(city_name,pr_name)
        b_shi = dict(Counter(biaoge_shi))
        # print(b_shi)
        if [key for key,value in b_shi.items() if value > 1] == []:
            print([key for key, value in b_shi.items() if value > 1])
            print(pr_name)
    # print(biaoge_sheng)
    b_sheng = dict(Counter(biaoge_sheng))
    # print(b_sheng)
    if {key:value for key,value in b_sheng.items()if value > 1} == {}:
        print({key:value for key,value in b_sheng.items()if value > 1})
    # print(dict(Counter(biaoge_sheng)))
        # print(len(b_sheng))
if __name__ == "__main__":
    shop_search()
