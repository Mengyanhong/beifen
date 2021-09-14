# import numpy as np
#
# a = [1, 1, 2,"3"]
# # 先转化为numpy.ndarray
# arr = np.array(a)
#
# # 是否全部为1,False
# # print((arr == 1).all())
#
# # 是否含有1,True
# print((arr == "3").any())
from random import random

import openpyxl
import requests

shopcategory = [{"name": "10", "value": "美食", "sub": [{"name": "112", "value": "小吃快餐"}, {"name": "110", "value": "火锅"},
                                                      {"name": "2714", "value": "水果生鲜"}, {"name": "508", "value": "烧烤"},
                                                      {"name": "118", "value": "其他美食"},
                                                      {"name": "34236", "value": "饮品店"}, {"name": "102", "value": "川菜"},
                                                      {"name": "117", "value": "面包/饮品"},
                                                      {"name": "33759", "value": "食品保健"},
                                                      {"name": "215", "value": "面馆"}, {"name": "116", "value": "西餐"},
                                                      {"name": "101", "value": "本帮江浙菜"},
                                                      {"name": "114", "value": "韩国料理"}, {"name": "132", "value": "咖啡厅"},
                                                      {"name": "106", "value": "东北菜"},
                                                      {"name": "25474", "value": "农家菜"}, {"name": "251", "value": "海鲜"},
                                                      {"name": "103", "value": "潮粤菜"}, {"name": "104", "value": "湘菜"},
                                                      {"name": "1783", "value": "家常菜"}, {"name": "113", "value": "日本菜"},
                                                      {"name": "34284", "value": "特色菜"},
                                                      {"name": "1338", "value": "私房菜"}, {"name": "219", "value": "小龙虾"},
                                                      {"name": "311", "value": "北京菜"}, {"name": "111", "value": "自助餐"},
                                                      {"name": "3243", "value": "新疆菜"},
                                                      {"name": "6743", "value": "云贵菜"},
                                                      {"name": "26481", "value": "西北菜"},
                                                      {"name": "250", "value": "创意菜"}, {"name": "115", "value": "东南亚菜"},
                                                      {"name": "1817", "value": "粉面馆"}, {"name": "246", "value": "湖北菜"},
                                                      {"name": "109", "value": "素食"}, {"name": "1959", "value": "粥粉面"},
                                                      {"name": "27686", "value": "海南菜"},
                                                      {"name": "34059", "value": "福建菜"},
                                                      {"name": "1080", "value": "豫菜"}, {"name": "107", "value": "台湾菜"},
                                                      {"name": "249", "value": "闽菜"}, {"name": "497", "value": "天津菜"}]}]
shop = list(shopcategory)
shop_list = shop[0]["sub"]
shop_len = len(shop[0]["sub"])

def shop_file():
    wb = openpyxl.Workbook()
    # wl = wb.create_sheet("1")
    ws = wb.active
    ws.cell(row=1, column=1).value = "name"
    ws.cell(row=1, column=2).value = "value"
    for i in range(shop_len):
        ws.cell(row=(i + 2), column=1).value = shop[0]["sub"][i]["name"]
        ws.cell(row=(i + 2), column=2).value = shop[0]["sub"][i]["value"]
    wb.save(r"C:/Users/admin/Desktop/店铺分类.xlsx")
    print("保存成功")

def shop_fenlei():
    for i in range(shop_len):
        url = "http://skb-test.weiwenjia.com/api_skb/v1/shop_search"
        headers = {
            'Authorization': 'Token token=2f5a36bebd4edc91e1ea3f5e803fc4c2',
            'crm_platform_type': 'lixiaoyun'
        }
        data = {
            "shopName": "",
            "hasUnfolded": 0,
            "hasSyncClue": 0,
            "page": 1,
            "pagesize": 10,
            "condition": {
                "cn": "composite",
                "cr": "MUST",
                "cv": [
                    {
                        "cn": "category",
                        "cv": {
                            "categoryL1": [],
                            "categoryL2": [
                                shop_list[i]["name"]
                            ]
                        },
                        "cr": "IN"
                    }
                ]
            }
        }
        response = requests.post(url, headers=headers, json=data)
        response = response.json()
        for j in range(10):
            res_catL1 = response["data"]["items"][0]["categoryL1"]
            res_catL2 = response["data"]["items"][j]["categoryL2"]
            if res_catL1 != "美食":
                print("搜索错误")
            if res_catL2 != shop_list[i]["value"]:
                print("二级行业搜索错误")
        # print(response)


if __name__ == "__main__":
    shop_file()
