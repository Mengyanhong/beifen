import json,sys
import openpyxl
print(sys.argv[0])
projectname = 'adv_search'
print(f"{sys.argv[0].split(projectname)[0]}{projectname}/涵盖证书类型.xlsx")
# print(f"{sys.argv[0].split(projectname)[0]}{projectname}/涵盖证书类型.xlsx")
text_file_path = r'C:\Users\admin\Desktop\配置文件.txt'
text_file = open(text_file_path,'r+')
file = text_file.read().encode('utf-8').decode('utf-8')

# for key,value in b.items():
#     if value in b.items():
#         print(1)
# b = json.loads(b,encoding='utf-8')
# # a = b.encode('iso-8859-1').decode('utf-8')
# print(b)
# print(eval(b)['adminDivision'])
# import json

# if isinstance(a, int):
#     print "a is int"
# else:
#     print "a is not int"
# if isinstance(b, list):
#     print
#     "b is list"
# else:
#     print
#     "b is not list"
# if isinstance(c, tuple):
#     print
#     "c is tuple"
# else:
#     print
#     "c is not tuple"
# if isinstance(b, dict):
#     print("d is dict")
# else:
#     print("d not dict")
# if isinstance(b, str):
#     print("d is str")
#
# else:
#     print("d not str")
def check_json_format(raw_msg):
    """
    用于判断一个字符串是否符合Json格式
    :param self:
    :return:
    """
    if isinstance(raw_msg, str): # 首先判断变量是否为字符串
        try:
            congig_text = json.loads(raw_msg)#, encoding='utf-8'
        except ValueError:
            return False
        return congig_text
    else:
         return False

def shop_file(sum_text,exe_text):
    wb = openpyxl.Workbook()
    # wl = wb.create_sheet("1") #打开表格索引为1的表格
    # ws = wb.active #打开默认表格
    # print(wb.get_sheet_names())
    # print(wb.sheetnames())
    wb.active.title = '涵盖证书类型' #更改默认表名
    # ws = wb.create_sheet(title = 'biaoge',index= 1 ) #创建新表
    # ws = wb.get_sheet_by_name('涵盖证书类型') #打开更改后的表格
    ws = wb['涵盖证书类型']
    # print(exe_text)
    ws.cell(row=1, column=1).value = "certL1Type_name"
    ws.cell(row=1, column=2).value = "certL1Type_value"
    ws.cell(row=1, column=3).value = "certL1Type_order"
    ws.cell(row=1, column=4).value = "certL2Type_name"
    ws.cell(row=1, column=5).value = "certL2Type_value"
    ws.cell(row=1, column=6).value = "certL2Type_order"
    sum = 0
    for vlue in range(sum_text):
        exe_text_sub_sum = exe_text[vlue]['sub']
        exe_text_sub_sum_list = len(exe_text[vlue]['sub'])
        for i in range(exe_text_sub_sum_list):
            ws.cell(row=(sum + 2), column=1).value = exe_text[vlue]['value']
            ws.cell(row=(sum + 2), column=2).value = exe_text[vlue]['name']
            ws.cell(row=(sum + 2), column=3).value = exe_text[vlue]['order']
            ws.cell(row=(sum + 2), column=4).value = exe_text_sub_sum[i]['value']
            ws.cell(row=(sum + 2), column=5).value = exe_text_sub_sum[i]['name']
            ws.cell(row=(sum + 2), column=6).value = exe_text[vlue]['order']
            sum += 1
    wb.save(f"{sys.argv[0].split(projectname)[0]}{projectname}/涵盖证书类型.xlsx")
    print("保存成功")
if __name__ == "__main__":
    # print (check_json_format("""{"a":1}"""))
    # print (check_json_format("""{'a':1}"""))
    # print (check_json_format({'a': 1}))
    sum_text = len(check_json_format(file)['certType'])
    print(sum_text)
    exe_text = check_json_format(file)['certType']
    print(exe_text)
    shop_file(sum_text,exe_text)
        # print(check_json_format(b)[list[vlue]],'\n')



    # print(check_json_format(b)['adminDivision'])
    #
    #     if check_json_format(b):
    #         print(check_json_format(b))






# print(b)